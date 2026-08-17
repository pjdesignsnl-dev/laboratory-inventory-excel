"""Structural inspection and integrity checks for the v0.1 macro-free workbook.

Usage:
  python scripts/inspect_workbook.py [workbook_path]

Emits:
  1. A human-readable inventory (sheets, tables, columns, named ranges,
     formulas, validation, protection) to evidence/workbook-inventory.txt
  2. Machine-readable CSV exports to evidence/
  3. A PASS/FAIL integrity report on stdout and evidence/non-vba-structural-report.txt

Checks implemented:
  - exact sheet set matches contract
  - exact Table names per sheet
  - exact column names per Table
  - primary keys nonblank + unique (fixture data)
  - barcode uniqueness (text-aware) and 7-digit format
  - FK references (ProductID->tblProducts, SupplierID->tblSuppliers,
    StorageLocationID->tblLocations, Transaction ContainerID/Barcode/ProductID)
  - controlled values in list (Status, TransactionType)
  - formula cells present for stock/reorder/expiry/dashboard
  - barcodes stored as text (leading zero preserved)
  - transactions append-only (no duplicate TransactionID, no blank rows in history)
  - stock formula correctness delegated to tests/formula tests (formulas lib)
"""
from __future__ import annotations

import csv
import os
import re
import sys

_TOOLS_PYLIB = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".tools", "pylib"
)
if os.path.isdir(_TOOLS_PYLIB) and _TOOLS_PYLIB not in sys.path:
    sys.path.insert(0, _TOOLS_PYLIB)

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from wb_schema import (
    SHEET_ORDER, TABLES, NAMED_RANGES, LIST_RANGES, SETTINGS_NAMES, COLUMN_NAMES,
    STATUSES, TRANSACTION_TYPES, PRODUCT_TYPES, CATEGORIES, LOCATION_TYPES,
    DISPOSAL_REASONS, TRANSACTION_REASONS, BOOLEANS,
)

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


def col_letter(n):
    return get_column_letter(n)


def cell_ref(ws, row, col):
    return f"{col_letter(col)}{row}"


def main():
    wb_path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
        "workbook", "LabInventory_v0.1.xlsx",
    )
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    ev = os.path.join(root, "evidence")
    os.makedirs(ev, exist_ok=True)

    wb = load_workbook(wb_path, data_only=False)
    report = []
    ok = True

    def chk(cond, msg):
        nonlocal ok
        status = "PASS" if cond else "FAIL"
        if not cond:
            ok = False
        report.append(f"[{status}] {msg}")
        return cond

    # ---------------- sheet set
    chk(wb.sheetnames == SHEET_ORDER,
        f"sheet set matches contract: {wb.sheetnames}")

    # ---------------- tables & columns
    for tname, spec in TABLES.items():
        ws = wb[spec["sheet"]]
        if tname not in ws.tables:
            chk(False, f"table {tname} missing on {spec['sheet']}")
            continue
        tab = ws.tables[tname]
        # expected column headers from spec
        start_col = spec.get("start_col", 1)
        # header row = top row of table ref
        c1, r1, c2, r2 = range_boundaries(tab.ref)
        headers = []
        for i in range(len(spec["columns"])):
            headers.append(ws.cell(row=r1, column=start_col + i).value)
        expected = [c[1] for c in spec["columns"]]
        chk(headers == expected, f"{tname} headers match contract ({len(expected)} cols)")

    # ---------------- contract YAML vs workbook reconciliation (Correction 1)
    # Zero workbook-contract drift: the YAML contract columns must equal the
    # workbook's actual table columns exactly (names, order, count).
    try:
        import yaml as _yaml
        contract_path = os.path.join(root, "schema", "workbook-contract.yaml")
        with open(contract_path, encoding="utf-8") as f:
            contract = _yaml.safe_load(f)
        yaml_tables = {t["name"]: t for t in contract.get("tables", [])}
        drift = []
        for tname, spec in TABLES.items():
            ws = wb[spec["sheet"]]
            if tname not in ws.tables:
                continue
            tab = ws.tables[tname]
            c1, r1, c2, r2 = range_boundaries(tab.ref)
            start_col = spec.get("start_col", 1)
            wb_cols = [ws.cell(row=r1, column=start_col + i).value for i in range(len(spec["columns"]))]
            # match by schema column name -> contract column name
            yt = yaml_tables.get(tname)
            if yt is None:
                drift.append(f"{tname}: missing from contract YAML")
                continue
            yaml_cols = [c["name"] for c in yt.get("columns", [])]
            schema_names = [c[0] for c in spec["columns"]]
            if schema_names != yaml_cols:
                drift.append(
                    f"{tname}: schema names {schema_names} != contract YAML names {yaml_cols}"
                )
        chk(not drift, f"workbook vs contract YAML zero drift (issues: {drift or 'none'})")
        # also verify helper columns are documented as calculated in the contract
        helper_issues = []
        for tname, spec in TABLES.items():
            calc = spec.get("calculated_columns", [])
            yt = yaml_tables.get(tname)
            if yt is None:
                continue
            yaml_by_name = {c["name"]: c for c in yt.get("columns", [])}
            for cn in calc:
                cdef = yaml_by_name.get(cn)
                if cdef is None:
                    helper_issues.append(f"{tname}.{cn}: missing from contract")
                elif not (cdef.get("calculated") or cdef.get("authoritative") is False):
                    helper_issues.append(f"{tname}.{cn}: not marked calculated/non-authoritative in contract")
        chk(not helper_issues,
            f"helper/calculated columns documented as non-authoritative in contract (issues: {helper_issues or 'none'})")
    except Exception as _e:  # pragma: no cover
        chk(False, f"contract YAML reconciliation could not run: {_e}")

    # ---------------- named ranges
    defined = set(wb.defined_names.keys())
    expected_names = set(NAMED_RANGES) | set(LIST_RANGES) | set(SETTINGS_NAMES) | set(COLUMN_NAMES)
    missing = expected_names - defined
    chk(not missing, f"all contract named ranges present (missing: {sorted(missing) or 'none'})")

    # ---------------- fixtures: tables data
    products = wb["Products"]
    containers = wb["Containers"]
    transactions = wb["Transactions"]
    suppliers = wb["Suppliers"]
    locations = wb["Locations"]

    def table_body(ws, tname):
        tab = ws.tables[tname]
        c1, r1, c2, r2 = range_boundaries(tab.ref)
        cols = TABLES[tname]["columns"]
        header = {cols[i][0]: ws.cell(row=r1, column=c1 + i).value for i in range(len(cols))}
        rows = []
        for r in range(r1 + 1, r2 + 1):
            row = {name: ws.cell(row=r, column=c1 + i).value for i, (name, _h, _w) in enumerate(cols)}
            if any(v is not None and v != "" for v in row.values()):
                rows.append(row)
        return rows

    prod_rows = table_body(products, "tblProducts")
    cont_rows = table_body(containers, "tblContainers")
    trans_rows = table_body(transactions, "tblTransactions")
    supp_rows = table_body(suppliers, "tblSuppliers")
    loc_rows = table_body(locations, "tblLocations")

    # PK uniqueness
    for tname, rows, key in [
        ("tblProducts", prod_rows, "ProductID"),
        ("tblContainers", cont_rows, "ContainerID"),
        ("tblTransactions", trans_rows, "TransactionID"),
        ("tblSuppliers", supp_rows, "SupplierID"),
        ("tblLocations", loc_rows, "StorageLocationID"),
    ]:
        vals = [r.get(key) for r in rows]
        nonblank = [v for v in vals if v not in (None, "")]
        chk(len(nonblank) == len(set(nonblank)), f"{tname}: {key} unique (n={len(nonblank)})")
        chk(len(nonblank) == len(vals), f"{tname}: {key} nonblank in all fixture rows")

    # barcode uniqueness + format + text storage
    barcodes = [r.get("Barcode") for r in cont_rows]
    chk(len(barcodes) == len(set(barcodes)), "tblContainers: Barcode unique")
    for bc in barcodes:
        if not (isinstance(bc, str) and re.fullmatch(r"\d{7}", bc)):
            chk(False, f"barcode {bc!r} not 7-digit text")
            break
    else:
        chk(True, "all barcodes are 7-digit text (leading zeros preserved)")

    # FK integrity
    prod_ids = {r["ProductID"] for r in prod_rows}
    supp_ids = {r["SupplierID"] for r in supp_rows}
    loc_ids = {r["StorageLocationID"] for r in loc_rows}
    cont_ids = {r["ContainerID"] for r in cont_rows}
    bad_prod_refs = {r["ProductID"] for r in cont_rows if r.get("ProductID") not in prod_ids}
    bad_supp_refs = {r["SupplierID"] for r in prod_rows if r.get("SupplierID") not in supp_ids}
    bad_loc_refs = {r["StorageLocationID"] for r in cont_rows if r.get("StorageLocationID") not in loc_ids}
    bad_txn_cont = {r["ContainerID"] for r in trans_rows if r.get("ContainerID") not in cont_ids}
    bad_txn_prod = {r["ProductID"] for r in trans_rows if r.get("ProductID") not in prod_ids}
    chk(not bad_prod_refs, f"Containers->Products FK valid (bad: {sorted(bad_prod_refs) or 'none'})")
    chk(not bad_supp_refs, f"Products->Suppliers FK valid (bad: {sorted(bad_supp_refs) or 'none'})")
    chk(not bad_loc_refs, f"Containers->Locations FK valid (bad: {sorted(bad_loc_refs) or 'none'})")
    chk(not bad_txn_cont, f"Transactions->Containers FK valid (bad: {sorted(bad_txn_cont) or 'none'})")
    chk(not bad_txn_prod, f"Transactions->Products FK valid (bad: {sorted(bad_txn_prod) or 'none'})")

    # barcodes in transactions must exist in containers
    cont_bc = {r["Barcode"] for r in cont_rows}
    bad_txn_bc = {r["Barcode"] for r in trans_rows if r.get("Barcode") not in cont_bc}
    chk(not bad_txn_bc, f"Transactions->Barcode FK valid (bad: {sorted(bad_txn_bc) or 'none'})")

    # controlled values
    bad_status = {r["Status"] for r in cont_rows if r.get("Status") not in STATUSES}
    bad_tt = {r["TransactionType"] for r in trans_rows if r.get("TransactionType") not in TRANSACTION_TYPES}
    bad_pt = {r["ProductType"] for r in prod_rows if r.get("ProductType") not in PRODUCT_TYPES}
    bad_cat = {r["Category"] for r in prod_rows if r.get("Category") not in CATEGORIES}
    chk(not bad_status, f"Container Status values valid (bad: {sorted(bad_status) or 'none'})")
    chk(not bad_tt, f"TransactionType values valid (bad: {sorted(bad_tt) or 'none'})")
    chk(not bad_pt, f"ProductType values valid (bad: {sorted(bad_pt) or 'none'})")
    chk(not bad_cat, f"Category values valid (bad: {sorted(bad_cat) or 'none'})")

    # expiry date >= received date where both present
    bad_dates = [r["ContainerID"] for r in cont_rows
                 if r.get("ExpiryDate") and r.get("DateReceived")
                 and r["ExpiryDate"] < r["DateReceived"]]
    chk(not bad_dates, f"ExpiryDate >= DateReceived where both set (bad: {bad_dates or 'none'})")

    # date cells must be real dates (not text) so Excel date comparisons work
    date_cols = ["ExpiryDate", "RetestDate", "DateReceived", "OpenedDate", "DisposalDate"]
    bad_date_types = []
    for r in cont_rows:
        for dc in date_cols:
            v = r.get(dc)
            if v is not None and not hasattr(v, "isoformat"):
                bad_date_types.append((r.get("ContainerID"), dc, type(v).__name__))
    chk(not bad_date_types,
        f"date cells are real dates (bad: {bad_date_types or 'none'})")

    # No Reserved anywhere in status list or fixtures (D-016)
    reserved_hits = []
    for r in cont_rows:
        if r.get("Status") == "Reserved":
            reserved_hits.append(r.get("ContainerID"))
    # status list table check
    settings_ws = wb["Settings"]
    for r in range(1, 60):
        if settings_ws.cell(row=r, column=1).value == "Reserved":
            reserved_hits.append("tblStatusList row %d" % r)
    chk(not reserved_hits, f"No Reserved in status list/fixtures (hits: {reserved_hits or 'none'}) (D-016)")

    # Dashboard must include frequently-used + by-location sections (D-019)
    dash = wb["Dashboard"]
    dash_texts = []
    for r in range(1, dash.max_row + 1):
        v = dash.cell(row=r, column=2).value
        if isinstance(v, str):
            dash_texts.append(v)
    has_freq = any("frequently used" in t.lower() for t in dash_texts)
    has_loc = any("storage location" in t.lower() for t in dash_texts)
    chk(has_freq, "Dashboard has 'most frequently used products' section (D-019)")
    chk(has_loc, "Dashboard has 'inventory by storage location' section (D-019)")

    # formula presence: Products helper columns, Dashboard stats, Scan lookups
    formula_checks = []
    # Products helper T/U present as formulas
    tab = products.tables["tblProducts"]
    c1, r1, c2, r2 = range_boundaries(tab.ref)
    helper_t = products.cell(row=r1 + 1, column=20).value
    helper_u = products.cell(row=r1 + 1, column=21).value
    formula_checks.append(("Products HelperAvailableStock formula", isinstance(helper_t, str) and helper_t.startswith("=")))
    formula_checks.append(("Products HelperStockClass formula", isinstance(helper_u, str) and helper_u.startswith("=")))
    # Dashboard stat cells
    dash = wb["Dashboard"]
    dash_has = any(isinstance(dash.cell(row=r, column=3).value, str)
                   and dash.cell(row=r, column=3).value.startswith("=")
                   for r in range(5, 20))
    formula_checks.append(("Dashboard has formula stats", dash_has))
    # Scan staging row
    scan = wb["Scan"]
    scan_has = any(isinstance(scan.cell(row=13, column=c).value, str)
                   and scan.cell(row=13, column=c).value.startswith("=")
                   for c in range(4, 17))
    formula_checks.append(("Scan staging row has formulas", scan_has))
    for msg, cond in formula_checks:
        chk(cond, msg)

    # transaction append-only: all TransactionIDs unique already checked; also
    # ensure no empty TransactionID in body
    chk(all(r.get("TransactionID") for r in trans_rows), "Transactions: all rows have TransactionID (append-only integrity)")

    # protection
    all_protected = all(wb[s].protection.sheet for s in SHEET_ORDER)
    chk(all_protected, "all sheets protected")
    chk(wb.security.lockStructure, "workbook structure protected")

    # ---------------- exports
    inv_path = os.path.join(ev, "workbook-inventory.txt")
    with open(inv_path, "w", encoding="utf-8") as f:
        f.write("=== Laboratory Inventory v0.1 — workbook inventory ===\n")
        f.write(f"Workbook: {os.path.abspath(wb_path)}\n\n")
        f.write("SHEETS:\n")
        for s in wb.sheetnames:
            f.write(f"  {s}  (tables: {', '.join(wb[s].tables.keys()) or 'none'})\n")
        f.write("\nTABLES AND COLUMNS:\n")
        for tname, spec in TABLES.items():
            ws = wb[spec["sheet"]]
            tab = ws.tables.get(tname)
            ref = tab.ref if tab else "MISSING"
            f.write(f"\n{tname} on {spec['sheet']} ref {ref}\n")
            for name, header, w in spec["columns"]:
                f.write(f"    {name:32} {header}\n")
        f.write("\nNAMED RANGES:\n")
        for rname in sorted(wb.defined_names.keys()):
            dn = wb.defined_names[rname]
            f.write(f"  {rname} -> {dn.attr_text}\n")
        f.write("\nFORMULAS (sampled, one per category):\n")
        f.write("  Products T (AvailableStock): {!r}\n".format(helper_t))
        f.write("  Products U (StockClass): {!r}\n".format(helper_u))
        f.write("  Dashboard B6: {!r}\n".format(dash.cell(row=6, column=3).value))
        f.write("\nVALIDATION RULES:\n")
        for ws in wb.worksheets:
            for dv in ws.data_validations.dataValidation:
                f.write(f"  {ws.title}: {dv.type} {dv.formula1} -> {dv.sqref}\n")
        f.write("\nCONDITIONAL FORMATTING:\n")
        for ws in wb.worksheets:
            for rng in ws.conditional_formatting:
                f.write(f"  {ws.title}: {rng.sqref}\n")
                for rule in rng.rules:
                    f.write(f"      {rule.type}: {rule.formula}\n")
        f.write("\nPROTECTION:\n")
        for s in SHEET_ORDER:
            ws = wb[s]
            f.write(f"  {s}: protected={ws.protection.sheet} password={ws.protection.password!r}\n")
        f.write(f"  workbook structure locked: {wb.security.lockStructure}\n")

    # formula inventory CSV
    formula_rows = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_rows.append((ws.title, cell.coordinate, cell.value))
    with open(os.path.join(ev, "formula-inventory.csv"), "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Sheet", "Cell", "Formula"])
        for row in formula_rows:
            w.writerow(row)

    # validation inventory CSV
    with open(os.path.join(ev, "validation-inventory.csv"), "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(["Sheet", "Type", "Formula1", "Sqref"])
        for ws in wb.worksheets:
            for dv in ws.data_validations.dataValidation:
                w.writerow([ws.title, dv.type, dv.formula1, str(dv.sqref)])

    # CSV exports
    for tname, spec in TABLES.items():
        ws = wb[spec["sheet"]]
        tab = ws.tables.get(tname)
        if not tab:
            continue
        c1, r1, c2, r2 = range_boundaries(tab.ref)
        cols = spec["columns"]
        csv_path = os.path.join(ev, f"table-{tname}.csv")
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow([c[1] for c in cols])
            for r in range(r1 + 1, r2 + 1):
                row = []
                for i in range(len(cols)):
                    v = ws.cell(row=r, column=c1 + i).value
                    if hasattr(v, "isoformat"):
                        v = v.isoformat()
                    row.append("" if v is None else v)
                if any(row):
                    w.writerow(row)

    # summary
    summary_path = os.path.join(ev, "non-vba-structural-report.txt")
    with open(summary_path, "w", encoding="utf-8") as f:
        f.write("NON-VBA STRUCTURAL INSPECTION REPORT\n")
        f.write(f"Workbook: {os.path.abspath(wb_path)}\n")
        f.write(f"Overall: {'PASS' if ok else 'FAIL'}\n\n")
        for line in report:
            f.write(line + "\n")

    print("\n".join(report))
    print(f"\nOVERALL: {'PASS' if ok else 'FAIL'}")
    return 0 if ok else 1


if __name__ == "__main__":
    sys.exit(main())
