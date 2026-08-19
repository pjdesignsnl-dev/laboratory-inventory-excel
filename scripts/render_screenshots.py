"""Render representative screenshots of the v0.1 workbook with Pillow.

Reads the workbook's actual cells (values, number formats, fills) and draws
PNG previews. This is a structural representation, NOT a desktop-Excel
screenshot; the evidence file labels it as such.

Run:  python scripts/render_screenshots.py
Outputs: evidence/screenshots/*.png
"""
from __future__ import annotations

import os
import sys

_TOOLS_PYLIB = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".tools", "pylib"
)
if os.path.isdir(_TOOLS_PYLIB) and _TOOLS_PYLIB not in sys.path:
    sys.path.insert(0, _TOOLS_PYLIB)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
WB = os.path.join(ROOT, "workbook", "LabInventory_v0.1.xlsx")
OUT = os.path.join(ROOT, "evidence", "screenshots")

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

CELL_W = 110
CELL_H = 22
HEADER_H = 30
FONT_PATH = r"C:\Windows\Fonts\arial.ttf"
try:
    FONT = ImageFont.truetype(FONT_PATH, 11)
    FONT_BOLD = ImageFont.truetype(FONT_PATH, 12)
    FONT_SMALL = ImageFont.truetype(FONT_PATH, 9)
except Exception:
    FONT = ImageFont.load_default()
    FONT_BOLD = ImageFont.load_default()
    FONT_SMALL = ImageFont.load_default()

HEX_FILL = {
    "1F4E79": (31, 78, 121), "E8F0FE": (232, 240, 254), "FFF3CD": (255, 243, 205),
    "D4EDDA": (212, 237, 218), "F8D7DA": (248, 215, 218), "EEEEEE": (238, 238, 238),
    "FFFDE7": (255, 253, 231), "FFFFFF": (255, 255, 255), "607D8B": (96, 125, 139),
}


def fill_rgb(cell):
    fill = cell.fill
    if fill and fill.fill_type == "solid" and fill.fgColor:
        rgb = fill.fgColor.rgb
        if rgb and isinstance(rgb, str) and len(rgb) >= 6:
            return HEX_FILL.get(rgb[-6:], (255, 255, 255))
    return (255, 255, 255)


def cell_text(cell):
    v = cell.value
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    return str(v)


def render_sheet(wb, sheet_name, max_rows, max_cols, out_name):
    ws = wb[sheet_name]
    img = Image.new("RGB", (max_cols * CELL_W + 2, max_rows * CELL_H + 2), (255, 255, 255))
    d = ImageDraw.Draw(img)
    for r in range(1, max_rows + 1):
        for c in range(1, max_cols + 1):
            cell = ws.cell(row=r, column=c)
            x0 = (c - 1) * CELL_W
            y0 = (r - 1) * CELL_H
            bg = fill_rgb(cell)
            d.rectangle([x0, y0, x0 + CELL_W, y0 + CELL_H], fill=bg, outline=(176, 190, 197))
            txt = cell_text(cell)
            if txt:
                font = FONT_SMALL if len(txt) > 28 else FONT
                d.text((x0 + 3, y0 + 3), txt[:42], fill=(20, 20, 20), font=font)
    os.makedirs(OUT, exist_ok=True)
    p = os.path.join(OUT, out_name)
    img.save(p)
    print("WROTE", p, img.size)


def main():
    wb = load_workbook(WB, data_only=False)
    # Dashboard: rows 1..40, cols 1..6
    render_sheet(wb, "Dashboard", 40, 6, "01-dashboard.png")
    # Scan: rows 1..32, cols 1..10
    render_sheet(wb, "Scan", 32, 10, "02-scan.png")
    # Receiving: rows 1..32, cols 1..9
    render_sheet(wb, "Receiving", 32, 9, "03-receiving.png")
    # Products: rows 1..10, cols 1..21
    render_sheet(wb, "Products", 10, 21, "04-products.png")
    # Containers: rows 1..24, cols 1..15
    render_sheet(wb, "Containers", 24, 15, "05-containers.png")
    # Transactions: rows 1..8, cols 1..16 (sample)
    render_sheet(wb, "Transactions", 8, 16, "06-transactions.png")


if __name__ == "__main__":
    main()
