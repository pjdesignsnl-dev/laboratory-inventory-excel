"""Build the macro-free Laboratory Inventory v0.1 workbook.

Run:  python scripts/build_workbook.py [output_path]
Produces workbook/LabInventory_v0.1.xlsx with nine sheets, nine Tables,
formulas, validation, conditional formatting, protection design, and
synthetic fixtures. No VBA is written or embedded.
"""
from __future__ import annotations

import os
import sys

# workspace-local third-party libs (openpyxl, formulas, pillow)
_TOOLS_PYLIB = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".tools", "pylib"
)
if os.path.isdir(_TOOLS_PYLIB) and _TOOLS_PYLIB not in sys.path:
    sys.path.insert(0, _TOOLS_PYLIB)

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from wb_schema import (
    SHEET_ORDER, TAB_COLORS, TABLES, NAMED_RANGES, LIST_RANGES, SETTINGS_NAMES,
    SETTINGS, STATUSES, TRANSACTION_TYPES, EXPIRY_CLASSES, PRODUCT_TYPES,
    CATEGORIES, LOCATION_TYPES, DISPOSAL_REASONS, TRANSACTION_REASONS, BOOLEANS,
    COLUMN_NAMES, WORKBOOK_VERSION, CONTRACT_VERSION,
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

# ------------------------------------------------------------------ styles
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F4E79")
SUB_FONT = Font(size=10, color="555555")
SECTION_FONT = Font(bold=True, size=12, color="1F4E79")
BODY_FONT = Font(size=10)
SMALL_FONT = Font(size=9, italic=True, color="777777")
THIN = Side(style="thin", color="B0BEC5")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
FILL_LIGHT = PatternFill("solid", fgColor="E8F0FE")
FILL_WARN = PatternFill("solid", fgColor="FFF3CD")
FILL_OK = PatternFill("solid", fgColor="D4EDDA")
FILL_BAD = PatternFill("solid", fgColor="F8D7DA")
FILL_GRAY = PatternFill("solid", fgColor="EEEEEE")
FILL_INPUT = PatternFill("solid", fgColor="FFFDE7")

NUM_FMT_DATE = "yyyy-mm-dd"
NUM_FMT_DATETIME = "yyyy-mm-dd hh:mm"
NUM_FMT_TEXT = "@"

# date anchors for synthetic fixtures (fixed so tests are deterministic;
# the workbook itself uses TODAY() for live classification)
BASE_DATE = "2026-08-17"


def _style_header(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX


def _title(ws, text, sub=None, row=1, ncols=10):
    ws.cell(row=row, column=1, value=text).font = TITLE_FONT
    if sub:
        ws.cell(row=row + 1, column=1, value=sub).font = SUB_FONT


def _set_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _box_region(ws, r1, c1, r2, c2, fill=None):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BOX
            if fill:
                cell.fill = fill


def build(output_path: str) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    for name in SHEET_ORDER:
        ws = wb.create_sheet(name)
        ws.sheet_properties.tabColor = TAB_COLORS.get(name, "808080")

    # ------------------------------------------------------------ Settings
    _build_settings(wb)

    # ------------------------------------------------------------ data tables
    _build_suppliers(wb)
    _build_locations(wb)
    _build_products(wb)
    _build_containers(wb)
    _build_transactions(wb)

    # ------------------------------------------------------------ Scan / Receiving / Dashboard
    _build_scan(wb)
    _build_receiving(wb)
    _build_dashboard(wb)

    # ------------------------------------------------------------ tables / names / validation / CF / protection
    _apply_tables(wb)
    _apply_named_ranges(wb)
    _apply_validation(wb)
    _apply_conditional_formatting(wb)
    _apply_protection(wb)

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    return output_path


# ------------------------------------------------------------------ Settings
def _build_settings(wb):
    ws = wb["Settings"]
    _title(ws, "Settings & Controlled Lists", "Constants, statuses, transaction types, expiry classes. Read-mostly — protect this sheet.", ncols=8)
    _set_widths(ws, [26, 16, 46, 4, 14, 14, 14, 14])

    ws.cell(row=4, column=1, value="tblSettings — configurable constants").font = SECTION_FONT
    hdr = 5
    for i, col in enumerate(TABLES["tblSettings"]["columns"], start=1):
        ws.cell(row=hdr, column=i, value=col[1])
    _style_header(ws, hdr, 3)
    # rows 6..12 for 7 settings
    for idx, (key, (val, desc)) in enumerate(SETTINGS.items()):
        r = 6 + idx
        ws.cell(row=r, column=1, value=key)
        ws.cell(row=r, column=2, value=val)
        ws.cell(row=r, column=3, value=desc)
    ws.cell(row=13, column=1, value="Settings table (7 keys): expiry bands 30/60/90 days, default location, default new status, scanner Enter suffix, version.").font = SMALL_FONT

    # status list (tblStatusList) — header text == column name (Excel refs)
    sl_row = 15
    ws.cell(row=sl_row, column=1, value="tblStatusList").font = SECTION_FONT
    ws.cell(row=sl_row + 1, column=1, value="StatusValue").font = HDR_FONT
    ws.cell(row=sl_row + 1, column=2, value="Label").font = HDR_FONT
    _style_header(ws, sl_row + 1, 2)
    status_labels = {
        "Available": "In storage, usable",
        "InUse": "Taken/opened, out of storage",
        "Expired": "Past expiry, unusable",
        "Damaged": "Physically unusable",
        "Disposed": "Removed from inventory",
        "Missing": "Cannot be located",
    }
    for i, s in enumerate(STATUSES):
        ws.cell(row=sl_row + 2 + i, column=1, value=s)
        ws.cell(row=sl_row + 2 + i, column=2, value=status_labels[s])

    # transaction type list
    tt_row = sl_row + 2 + len(STATUSES) + 1
    ws.cell(row=tt_row, column=1, value="tblTransactionTypeList").font = SECTION_FONT
    ws.cell(row=tt_row + 1, column=1, value="TransactionTypeValue").font = HDR_FONT
    ws.cell(row=tt_row + 1, column=2, value="Label").font = HDR_FONT
    _style_header(ws, tt_row + 1, 2)
    tt_labels = {
        "Receive": "Create/activate a container",
        "TakeOpen": "Take from storage / open",
        "Return": "Return to storage",
        "Transfer": "Change storage location",
        "Dispose": "Remove from inventory",
        "MarkExpired": "Mark as expired",
        "MarkDamaged": "Mark as damaged",
        "MarkMissing": "Mark as missing/lost",
        "Adjustment": "Correction/reversal (compensating)",
    }
    for i, t in enumerate(TRANSACTION_TYPES):
        ws.cell(row=tt_row + 2 + i, column=1, value=t)
        ws.cell(row=tt_row + 2 + i, column=2, value=tt_labels[t])

    # expiry class list
    ec_row = tt_row + 2 + len(TRANSACTION_TYPES) + 1
    ws.cell(row=ec_row, column=1, value="tblExpiryClassList").font = SECTION_FONT
    ws.cell(row=ec_row + 1, column=1, value="ExpiryClassValue").font = HDR_FONT
    ws.cell(row=ec_row + 1, column=2, value="Label").font = HDR_FONT
    _style_header(ws, ec_row + 1, 2)
    ec_labels = {
        "Expired": "Date < today",
        "ExpiringSoon": "Within warning band (30/60/90)",
        "Valid": "Beyond warning band",
        "NoExpiry": "No expiry date set",
        "Invalid": "Date error",
    }
    for i, e in enumerate(EXPIRY_CLASSES):
        ws.cell(row=ec_row + 2 + i, column=1, value=e)
        ws.cell(row=ec_row + 2 + i, column=2, value=ec_labels[e])

    # inline list sources (named ranges lst*) — written to column H to match
    # the schema LIST_RANGES addresses (H4:H6, H8:H13, H15:H20, H22:H26, H28:H33, H35:H36)
    lst_title = ec_row + 2 + len(EXPIRY_CLASSES) + 2   # 45
    ws.cell(row=lst_title, column=1, value="Inline list sources (named ranges lst*):").font = SECTION_FONT
    ws.cell(row=lst_title, column=7, value="ProductType").font = HDR_FONT
    ws.cell(row=lst_title, column=8, value="Category").font = HDR_FONT
    for i, v in enumerate(PRODUCT_TYPES):
        ws.cell(row=4 + i, column=8, value=v)  # H4:H6
    for i, v in enumerate(CATEGORIES):
        ws.cell(row=8 + i, column=8, value=v)  # H8:H13
    ws.cell(row=14, column=7, value="LocationType").font = HDR_FONT
    for i, v in enumerate(LOCATION_TYPES):
        ws.cell(row=15 + i, column=8, value=v)  # H15:H20
    ws.cell(row=21, column=7, value="DisposalReason").font = HDR_FONT
    for i, v in enumerate(DISPOSAL_REASONS):
        ws.cell(row=22 + i, column=8, value=v)  # H22:H26
    ws.cell(row=27, column=7, value="TransactionReason").font = HDR_FONT
    for i, v in enumerate(TRANSACTION_REASONS):
        ws.cell(row=28 + i, column=8, value=v)  # H28:H33
    ws.cell(row=34, column=7, value="Bool").font = HDR_FONT
    for i, v in enumerate(BOOLEANS):
        ws.cell(row=35 + i, column=8, value=v)  # H35:H36

    ws.cell(row=lst_title + 1, column=1, value="Protection: this sheet is locked (read-mostly). Change constants only after unprotecting.").font = SMALL_FONT


# ------------------------------------------------------------------ data tables
def _build_suppliers(wb):
    ws = wb["Suppliers"]
    spec = TABLES["tblSuppliers"]
    cols = spec["columns"]
    _title(ws, "tblSuppliers", "Supplier catalogue. SupplierID references from tblProducts.", ncols=len(cols))
    _set_widths(ws, [c[2] for c in cols])
    hdr = 4
    for i, (name, header, _w) in enumerate(cols, start=1):
        ws.cell(row=hdr, column=i, value=header)
    _style_header(ws, hdr, len(cols))
    rows = [
        ("S000001", "LabSupply Co.", "A. Rivera", "orders@labsupply.example", "+1 555 0101", "labsupply.example", "123 Industrial Pkwy, Springfield", "Primary general supplier"),
        ("S000002", "ChemCore Ltd.", "B. Novak", "sales@chemcore.example", "+1 555 0102", "chemcore.example", "45 Analytical Way, Brookfield", "Chemicals/reagents"),
        ("S000003", "TipTech GmbH", "C. Meier", "info@tiptech.example", "+49 30 555 0103", "tiptech.example", "9 Laborenstrasse, Berlin", "Pipette tips and plastics"),
    ]
    for i, row in enumerate(rows, start=hdr + 1):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)


def _build_locations(wb):
    ws = wb["Locations"]
    spec = TABLES["tblLocations"]
    cols = spec["columns"]
    _title(ws, "tblLocations", "Storage locations. StorageLocationID references from tblContainers.", ncols=len(cols))
    _set_widths(ws, [c[2] for c in cols])
    hdr = 4
    for i, (name, header, _w) in enumerate(cols, start=1):
        ws.cell(row=hdr, column=i, value=header)
    _style_header(ws, hdr, len(cols))
    rows = [
        ("LOC0001", "Chemical Cabinet 1", "Cabinet", "Flammables", True),
        ("LOC0002", "Chemical Cabinet 2", "Cabinet", "Acids", True),
        ("LOC0003", "Fridge 1", "Fridge", "Cold reagents", True),
        ("LOC0004", "Plastics Shelf A", "Shelf", "Pipette tips, tubes, gloves", True),
        ("LOC0005", "Plastics Shelf B", "Shelf", "Filters, plates", True),
        ("LOC0006", "Freezer -20C", "Freezer", "Frozen reagents", True),
    ]
    for i, row in enumerate(rows, start=hdr + 1):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)


def _build_products(wb):
    ws = wb["Products"]
    spec = TABLES["tblProducts"]
    cols = spec["columns"]
    _title(ws, "tblProducts", "Product catalogue. Stock/reorder are formula helper columns (protected).", ncols=len(cols))
    _set_widths(ws, [c[2] for c in cols])
    hdr = 4
    for i, (name, header, _w) in enumerate(cols, start=1):
        ws.cell(row=hdr, column=i, value=header)
    _style_header(ws, hdr, len(cols))
    rows = [
        # id, name, type, category, mfr, catno, cas, conc, grade, stddesc, supplier, storage, hazard, sds, min, target, reorder, active, notes
        ("P000001", "Pipette Tips 200 uL (96/box)", "Consumable", "Pipette Tips", "TipTech GmbH", "TT-200-96", None, None, None, "Box of 96 tips", "S000003", "Ambient", None, "SDS-001", 3, 8, 5, True, "Worked example 1"),
        ("P000002", "Laboratory Tubes 15 mL (100/box)", "Consumable", "Tubes", "LabSupply Co.", "LS-15-100", None, None, None, "Box of 100 tubes", "S000001", "Ambient", None, "SDS-002", 2, 6, 4, True, "Worked example 2"),
        ("P000003", "Ethanol 96% (500 mL)", "Chemical", "Solvent", "ChemCore Ltd.", "CC-ET-500", "64-17-5", "96%", "Analytical", "Bottle of 500 mL", "S000002", "Flammables cabinet", "Flammable", "SDS-ETOH", 2, 4, 3, True, "Worked example 3"),
        ("P000004", "Methanol 99.8% (1 L)", "Chemical", "Solvent", "ChemCore Ltd.", "CC-ME-1000", "67-56-1", "99.8%", "HPLC", "Bottle of 1 L", "S000002", "Flammables cabinet", "Flammable", "SDS-MEOH", 1, 3, 2, True, "Worked example 4"),
        ("P000005", "Reagent A (Tris buffer, 500 mL)", "Reagent", "Reagent", "LabSupply Co.", "LS-TRIS-500", "77-86-1", "1 M", "Molecular biology", "Bottle of 500 mL", "S000001", "Fridge 1", "Irritant", "SDS-TRIS", 1, 4, 2, True, "Worked example 5; expiring reagent"),
        ("P000006", "Nitrile Gloves M (100/box)", "Consumable", "Consumable", "LabSupply Co.", "LS-NG-M100", None, None, None, "Box of 100 gloves", "S000001", "Ambient", None, None, 4, 10, 6, True, "Consumable family"),
    ]
    for i, row in enumerate(rows, start=hdr + 1):
        for j, v in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=v)
    # helper formula columns (protected): AvailableStock, StockClass
    n = len(rows)
    for k in range(n):
        r = hdr + 1 + k
        pid = ws.cell(row=r, column=1).value
        # AvailableStock (D-018): Status=Available AND (ExpiryDate blank OR
        # ExpiryDate >= TODAY()). Implemented as: available minus
        # available-but-expired-by-date (COUNTIFS cannot express OR directly;
        # subtraction is 2021-compatible and auditable).
        ws.cell(row=r, column=20, value=(
            f'=COUNTIFS(tblContainers[ProductID],A{r},tblContainers[Status],"Available")'
            f'-COUNTIFS(tblContainers[ProductID],A{r},tblContainers[Status],"Available",'
            f'tblContainers[ExpiryDate],"<"&TODAY())'
        ))
        # StockClass
        ws.cell(row=r, column=21, value=f'=IF(T{r}=0,"OutOfStock",IF(T{r}<O{r},"Reorder",IF(T{r}<P{r},"Low","OK")))')
    # header for helper columns
    for i, (name, header, _w) in enumerate(cols, start=1):
        pass
    # style helper header cells
    ws.cell(row=hdr, column=20, value=cols[19][1])
    ws.cell(row=hdr, column=21, value=cols[20][1])
    ws.cell(row=hdr, column=20).font = HDR_FONT
    ws.cell(row=hdr, column=20).fill = PatternFill("solid", fgColor="607D8B")
    ws.cell(row=hdr, column=21).font = HDR_FONT
    ws.cell(row=hdr, column=21).fill = PatternFill("solid", fgColor="607D8B")


def _build_containers(wb):
    ws = wb["Containers"]
    spec = TABLES["tblContainers"]
    cols = spec["columns"]
    _title(ws, "tblContainers", "Every physical container/package = one row. Barcode unique. Stock = count of Available.", ncols=len(cols))
    _set_widths(ws, [c[2] for c in cols])
    hdr = 4
    for i, (name, header, _w) in enumerate(cols, start=1):
        ws.cell(row=hdr, column=i, value=header)
    _style_header(ws, hdr, len(cols))
    # synthetic containers (barcodes unique, text, leading zeros preserved)
    rows = [
        # cid, barcode, pid, lot, expiry, retest, received, loc, status, opened, disposed, reason, notes
        ("C000001", "0000001", "P000001", "LOT0001", None, None, "2026-01-10", "LOC0004", "Available", None, None, None, "Tips box 1"),
        ("C000002", "0000002", "P000001", "LOT0001", None, None, "2026-01-10", "LOC0004", "Available", None, None, None, "Tips box 2"),
        ("C000003", "0000003", "P000001", "LOT0002", None, None, "2026-03-05", "LOC0004", "Available", "2026-03-06", None, None, "Tips box 3 (opened, returned)"),
        ("C000004", "0000004", "P000002", "LOT0010", None, None, "2026-02-20", "LOC0004", "Available", None, None, None, "Tubes box 1"),
        ("C000005", "0000005", "P000002", "LOT0010", None, None, "2026-02-20", "LOC0004", "Available", None, None, None, "Tubes box 2"),
        ("C000006", "0000006", "P000002", "LOT0011", None, None, "2026-05-11", "LOC0005", "InUse", "2026-05-12", None, None, "Tubes box 3 in use"),
        ("C000007", "0000007", "P000003", "LOT0020", "2027-04-15", None, "2025-11-01", "LOC0001", "Available", None, None, None, "Ethanol bottle A"),
        ("C000008", "0000008", "P000003", "LOT0020", "2027-04-15", None, "2025-11-01", "LOC0001", "Available", None, None, None, "Ethanol bottle B"),
        ("C000009", "0000009", "P000003", "LOT0021", "2026-10-31", None, "2026-06-20", "LOC0001", "Available", "2026-07-01", None, None, "Ethanol bottle C (opened, returned)"),
        ("C000010", "0000010", "P000004", "LOT0030", "2026-12-31", None, "2026-04-18", "LOC0001", "Available", None, None, None, "Methanol bottle A"),
        ("C000011", "0000011", "P000004", "LOT0030", "2026-12-31", None, "2026-04-18", "LOC0002", "InUse", "2026-06-10", None, None, "Methanol bottle B in use"),
        ("C000012", "0000012", "P000005", "LOT0040", "2026-08-20", "2026-08-01", "2026-05-01", "LOC0003", "Available", None, None, None, "Tris reagent (expiring soon, band 30)"),
        ("C000013", "0000013", "P000005", "LOT0041", "2026-09-05", None, "2026-05-15", "LOC0003", "Available", None, None, None, "Tris reagent (band 60)"),
        ("C000014", "0000014", "P000005", "LOT0042", "2026-07-01", None, "2026-01-20", "LOC0003", "Expired", None, "2026-07-02", "Expired", "Tris reagent expired"),
        ("C000015", "0000015", "P000006", "LOT0050", None, None, "2026-04-01", "LOC0004", "Available", None, None, None, "Gloves box 1"),
        ("C000016", "0000016", "P000006", "LOT0050", None, None, "2026-04-01", "LOC0004", "Available", None, None, None, "Gloves box 2"),
        ("C000017", "0000017", "P000006", "LOT0051", None, None, "2026-07-10", "LOC0004", "Disposed", None, "2026-07-11", "Used Up", "Gloves box 3 disposed"),
        ("C000018", "0000018", "P000003", "LOT0022", "2027-06-30", None, "2026-08-10", "LOC0001", "Available", None, None, None, "Ethanol bottle D"),
        ("C000019", "0000019", "P000002", "LOT0012", None, None, "2026-08-12", "LOC0004", "Available", None, None, None, "Tubes box 4 (Available, no expiry)"),
        ("C000020", "0000020", "P000001", "LOT0003", None, None, "2026-08-15", "LOC0004", "Missing", None, None, "Missing", "Tips box 4 missing"),
        ("C000021", "0000021", "P000005", "LOT0043", "2026-08-10", None, "2026-07-15", "LOC0003", "Available", None, None, None, "Tris reagent expired-by-date but NOT yet MarkExpired (D-018 boundary)"),
    ]
    for i, row in enumerate(rows, start=hdr + 1):
        for j, v in enumerate(row, start=1):
            if j in (5, 6, 7, 10, 11) and isinstance(v, str) and v:
                # store REAL dates (openpyxl keeps strings as text otherwise,
                # which would break date comparisons in Excel)
                from datetime import date as _date
                v = _date.fromisoformat(v)
            cell = ws.cell(row=i, column=j, value=v)
            if j == 2:  # barcode text
                ws.cell(row=i, column=j).number_format = NUM_FMT_TEXT
            if j in (5, 6, 7, 10, 11):  # dates
                if v:
                    ws.cell(row=i, column=j).number_format = NUM_FMT_DATE
    # helper formula columns
    n = len(rows)
    for k in range(n):
        r = hdr + 1 + k
        ws.cell(row=r, column=14, value=f'=VALUE(SUBSTITUTE(A{r},"C",""))')
        ws.cell(row=r, column=15, value=f'=VALUE(B{r})')
    ws.cell(row=hdr, column=14, value="HelperContainerNum")
    ws.cell(row=hdr, column=15, value="HelperBarcodeNum")
    for c in (14, 15):
        ws.cell(row=hdr, column=c).font = HDR_FONT
        ws.cell(row=hdr, column=c).fill = PatternFill("solid", fgColor="607D8B")


def _build_transactions(wb):
    ws = wb["Transactions"]
    spec = TABLES["tblTransactions"]
    cols = spec["columns"]
    _title(ws, "tblTransactions", "APPEND-ONLY. Never delete or overwrite rows. Corrections use Adjustment transactions.", ncols=len(cols))
    _set_widths(ws, [c[2] for c in cols])
    hdr = 4
    for i, (name, header, _w) in enumerate(cols, start=1):
        ws.cell(row=hdr, column=i, value=header)
    _style_header(ws, hdr, len(cols))
    # synthetic transaction history: complete per-container lifecycles
    # (tid, ts, op, barcode, cid, pid, pname, type, prev, new, prevloc, newloc, lot, reason, ref, notes)
    P = {
        "P000001": "Pipette Tips 200 uL (96/box)",
        "P000002": "Laboratory Tubes 15 mL (100/box)",
        "P000003": "Ethanol 96% (500 mL)",
        "P000004": "Methanol 99.8% (1 L)",
        "P000005": "Reagent A (Tris buffer, 500 mL)",
        "P000006": "Nitrile Gloves M (100/box)",
    }
    rows = [
        ("T00000001", "2026-01-10 09:00", "LAB-USER", "0000001", "C000001", "P000001", P["P000001"], "Receive", "(none)", "Available", "(none)", "LOC0004", "LOT0001", None, "PO-1001", "Tips box 1 received"),
        ("T00000002", "2026-01-10 09:05", "LAB-USER", "0000002", "C000002", "P000001", P["P000001"], "Receive", "(none)", "Available", "(none)", "LOC0004", "LOT0001", None, "PO-1001", "Tips box 2 received"),
        ("T00000003", "2026-03-05 10:00", "LAB-USER", "0000003", "C000003", "P000001", P["P000001"], "Receive", "(none)", "Available", "(none)", "LOC0004", "LOT0002", None, "PO-1002", "Tips box 3 received"),
        ("T00000004", "2026-03-06 14:20", "LAB-USER", "0000003", "C000003", "P000001", P["P000001"], "TakeOpen", "Available", "InUse", "LOC0004", "(none)", "LOT0002", None, None, "Tips box 3 opened"),
        ("T00000005", "2026-03-07 09:30", "LAB-USER", "0000003", "C000003", "P000001", P["P000001"], "Return", "InUse", "Available", "(none)", "LOC0004", "LOT0002", None, None, "Tips box 3 returned (opened)"),
        ("T00000006", "2026-02-20 11:00", "LAB-USER", "0000004", "C000004", "P000002", P["P000002"], "Receive", "(none)", "Available", "(none)", "LOC0004", "LOT0010", None, "PO-1003", "Tubes box 1 received"),
        ("T00000007", "2026-02-20 11:05", "LAB-USER", "0000005", "C000005", "P000002", P["P000002"], "Receive", "(none)", "Available", "(none)", "LOC0004", "LOT0010", None, "PO-1003", "Tubes box 2 received"),
        ("T00000008", "2026-05-11 13:00", "LAB-USER", "0000006", "C000006", "P000002", P["P000002"], "Receive", "(none)", "Available", "(none)", "LOC0005", "LOT0011", None, "PO-1004", "Tubes box 3 received"),
        ("T00000009", "2026-05-12 08:45", "LAB-USER", "0000006", "C000006", "P000002", P["P000002"], "TakeOpen", "Available", "InUse", "LOC0005", "(none)", "LOT0011", None, None, "Tubes box 3 in use"),
        ("T00000010", "2025-11-01 10:30", "LAB-USER", "0000007", "C000007", "P000003", P["P000003"], "Receive", "(none)", "Available", "(none)", "LOC0001", "LOT0020", None, "PO-1005", "Ethanol A received"),
        ("T00000011", "2025-11-01 10:32", "LAB-USER", "0000008", "C000008", "P000003", P["P000003"], "Receive", "(none)", "Available", "(none)", "LOC0001", "LOT0020", None, "PO-1005", "Ethanol B received"),
        ("T00000012", "2026-06-20 09:15", "LAB-USER", "0000009", "C000009", "P000003", P["P000003"], "Receive", "(none)", "Available", "(none)", "LOC0001", "LOT0021", None, "PO-1006", "Ethanol C received"),
        ("T00000013", "2026-07-01 15:10", "LAB-USER", "0000009", "C000009", "P000003", P["P000003"], "TakeOpen", "Available", "InUse", "LOC0001", "(none)", "LOT0021", None, None, "Ethanol C opened"),
        ("T00000014", "2026-07-02 10:00", "LAB-USER", "0000009", "C000009", "P000003", P["P000003"], "Return", "InUse", "Available", "(none)", "LOC0001", "LOT0021", None, None, "Ethanol C returned (opened)"),
        ("T00000015", "2026-04-18 14:00", "LAB-USER", "0000010", "C000010", "P000004", P["P000004"], "Receive", "(none)", "Available", "(none)", "LOC0001", "LOT0030", None, "PO-1007", "Methanol A received"),
        ("T00000016", "2026-04-18 14:05", "LAB-USER", "0000011", "C000011", "P000004", P["P000004"], "Receive", "(none)", "Available", "(none)", "LOC0002", "LOT0030", None, "PO-1007", "Methanol B received"),
        ("T00000017", "2026-06-10 11:30", "LAB-USER", "0000011", "C000011", "P000004", P["P000004"], "TakeOpen", "Available", "InUse", "LOC0002", "(none)", "LOT0030", None, None, "Methanol B in use"),
        ("T00000018", "2026-05-01 09:00", "LAB-USER", "0000012", "C000012", "P000005", P["P000005"], "Receive", "(none)", "Available", "(none)", "LOC0003", "LOT0040", None, "PO-1008", "Tris reagent received"),
        ("T00000019", "2026-05-15 09:30", "LAB-USER", "0000013", "C000013", "P000005", P["P000005"], "Receive", "(none)", "Available", "(none)", "LOC0003", "LOT0041", None, "PO-1009", "Tris reagent 2 received"),
        ("T00000020", "2026-01-20 10:00", "LAB-USER", "0000014", "C000014", "P000005", P["P000005"], "Receive", "(none)", "Available", "(none)", "LOC0003", "LOT0042", None, "PO-1010", "Tris reagent 3 received"),
        ("T00000021", "2026-07-02 08:00", "LAB-USER", "0000014", "C000014", "P000005", P["P000005"], "MarkExpired", "Available", "Expired", "LOC0003", "LOC0003", "LOT0042", "Expired", None, "Tris reagent 3 expired"),
        ("T00000022", "2026-04-01 12:00", "LAB-USER", "0000015", "C000015", "P000006", P["P000006"], "Receive", "(none)", "Available", "(none)", "LOC0004", "LOT0050", None, "PO-1011", "Gloves box 1 received"),
        ("T00000023", "2026-04-01 12:02", "LAB-USER", "0000016", "C000016", "P000006", P["P000006"], "Receive", "(none)", "Available", "(none)", "LOC0004", "LOT0050", None, "PO-1011", "Gloves box 2 received"),
        ("T00000024", "2026-07-10 16:00", "LAB-USER", "0000017", "C000017", "P000006", P["P000006"], "Receive", "(none)", "Available", "(none)", "LOC0004", "LOT0051", None, "PO-1012", "Gloves box 3 received"),
        ("T00000025", "2026-07-11 09:20", "LAB-USER", "0000017", "C000017", "P000006", P["P000006"], "Dispose", "Available", "Disposed", "LOC0004", "(none)", "LOT0051", "Used Up", None, "Gloves box 3 used up"),
        ("T00000026", "2026-08-10 10:10", "LAB-USER", "0000018", "C000018", "P000003", P["P000003"], "Receive", "(none)", "Available", "(none)", "LOC0001", "LOT0022", None, "PO-1013", "Ethanol D received"),
        ("T00000027", "2026-08-12 11:00", "LAB-USER", "0000019", "C000019", "P000002", P["P000002"], "Receive", "(none)", "Available", "(none)", "LOC0004", "LOT0012", None, "PO-1014", "Tubes box 4 received"),
        ("T00000028", "2026-08-13 13:45", "LAB-USER", "0000019", "C000019", "P000002", P["P000002"], "Transfer", "Available", "Available", "LOC0004", "LOC0005", "LOT0012", None, None, "Tubes box 4 transferred"),
        ("T00000029", "2026-07-15 09:00", "LAB-USER", "0000021", "C000021", "P000005", P["P000005"], "Receive", "(none)", "Available", "(none)", "LOC0003", "LOT0043", None, "PO-1016", "Tris reagent received (expires 2026-08-10; not yet MarkExpired)"),
        ("T00000030", "2026-08-15 14:00", "LAB-USER", "0000020", "C000020", "P000001", P["P000001"], "Receive", "(none)", "Available", "(none)", "LOC0004", "LOT0003", None, "PO-1015", "Tips box 4 received"),
        ("T00000031", "2026-08-16 10:00", "LAB-USER", "0000020", "C000020", "P000001", P["P000001"], "MarkMissing", "Available", "Missing", "LOC0004", "(none)", "LOT0003", "Missing", None, "Tips box 4 missing"),
    ]
    for i, row in enumerate(rows, start=hdr + 1):
        for j, v in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j, value=v)
            if j == 2:
                cell.number_format = NUM_FMT_DATETIME
            if j == 9 or j == 10:
                pass
    # timestamp as datetime cells: convert date strings to datetime
    from datetime import datetime
    for k, row in enumerate(rows, start=hdr + 1):
        ts = row[1]
        ws.cell(row=k, column=2, value=datetime.strptime(ts, "%Y-%m-%d %H:%M"))
        ws.cell(row=k, column=2).number_format = NUM_FMT_DATETIME


def _build_scan(wb):
    ws = wb["Scan"]
    _title(ws, "Scan", "Scan or type a barcode (7 digits) then Enter. Lookup and validation are formula-driven — no VBA.", ncols=10)
    _set_widths(ws, [6, 16, 3, 18, 26, 14, 14, 14, 14, 12])

    ws.cell(row=4, column=2, value="Scan").font = SECTION_FONT
    ws.cell(row=6, column=2, value="Scan / type barcode:").font = BODY_FONT
    input_cell = ws.cell(row=7, column=4, value="")
    input_cell.fill = FILL_INPUT
    input_cell.font = Font(size=14, bold=True)
    input_cell.alignment = CENTER
    input_cell.border = BOX
    ws.cell(row=9, column=2, value="Status:").font = BODY_FONT
    ws.cell(row=9, column=4, value='=IF($D$7="","EMPTY",IF(ISNA(MATCH($D$7,tblContainers[Barcode],0)),"UNKNOWN",IF(COUNTIF(tblContainers[Barcode],$D$7)>1,"DUPLICATE","FOUND")))')
    ws.cell(row=9, column=4).font = Font(bold=True, size=12)

    # lookup result staging table (tblScanResults) header row 11
    ws.cell(row=11, column=4, value="Lookup result (formula staging)").font = SECTION_FONT
    scan_cols = TABLES["tblScanResults"]["columns"]
    hdr = 12
    for i, (name, header, _w) in enumerate(scan_cols, start=1):
        ws.cell(row=hdr, column=3 + i, value=header)
    _style_header(ws, hdr, len(scan_cols), start_col=4)
    # formula row 13
    f = {}
    # column letters: D=4..P=16
    f["Barcode"] = '=IF($D$7="","",$D$7)'
    f["ContainerID"] = '=IF($D$7="","",IF(ISNA(MATCH($D$7,tblContainers[Barcode],0)),"",INDEX(tblContainers[ContainerID],MATCH($D$7,tblContainers[Barcode],0))))'
    f["ProductID"] = '=IF($D$7="","",IF(ISNA(MATCH($D$7,tblContainers[Barcode],0)),"",INDEX(tblContainers[ProductID],MATCH($D$7,tblContainers[Barcode],0))))'
    f["ProductName"] = '=IF($D13="","",IF(ISNA(MATCH($F13,tblProducts[ProductID],0)),"",INDEX(tblProducts[ProductName],MATCH($F13,tblProducts[ProductID],0))))'
    f["BatchLotNumber"] = '=IF($D$7="","",IF(ISNA(MATCH($D$7,tblContainers[Barcode],0)),"",INDEX(tblContainers[BatchLotNumber],MATCH($D$7,tblContainers[Barcode],0))))'
    f["ExpiryDate"] = '=IF($D$7="","",IF(ISNA(MATCH($D$7,tblContainers[Barcode],0)),"",INDEX(tblContainers[ExpiryDate],MATCH($D$7,tblContainers[Barcode],0))))'
    f["StorageLocationID"] = '=IF($D$7="","",IF(ISNA(MATCH($D$7,tblContainers[Barcode],0)),"",INDEX(tblContainers[StorageLocationID],MATCH($D$7,tblContainers[Barcode],0))))'
    f["LocationName"] = '=IF($J13="","",IF(ISNA(MATCH($J13,tblLocations[StorageLocationID],0)),"",INDEX(tblLocations[LocationName],MATCH($J13,tblLocations[StorageLocationID],0))))'
    f["Status"] = '=IF($D$7="","",IF(ISNA(MATCH($D$7,tblContainers[Barcode],0)),"",INDEX(tblContainers[Status],MATCH($D$7,tblContainers[Barcode],0))))'
    f["OpenedDate"] = '=IF($D$7="","",IF(ISNA(MATCH($D$7,tblContainers[Barcode],0)),"",INDEX(tblContainers[OpenedDate],MATCH($D$7,tblContainers[Barcode],0))))'
    f["ExpiryClass"] = '=IF($I13="","NoExpiry",IF($I13<TODAY(),"Expired",IF($I13<=TODAY()+ExpiryWarningDays30,"ExpiringSoon",IF($I13<=TODAY()+ExpiryWarningDays60,"ExpiringSoon",IF($I13<=TODAY()+ExpiryWarningDays90,"ExpiringSoon","Valid")))))'
    f["DuplicateFlag"] = '=IF($D$7="","",IF(COUNTIF(tblContainers[Barcode],$D$7)>1,"DUPLICATE",""))'
    f["LookupState"] = '=IF($D$7="","EMPTY",IF(ISNA(MATCH($D$7,tblContainers[Barcode],0)),"UNKNOWN",IF(COUNTIF(tblContainers[Barcode],$D$7)>1,"DUPLICATE","FOUND")))'
    cols = scan_cols
    for i, (name, header, _w) in enumerate(cols, start=1):
        ws.cell(row=13, column=3 + i, value=f.get(name, ""))
        ws.cell(row=13, column=3 + i).border = BOX
        ws.cell(row=13, column=3 + i).alignment = CENTER

    # detail card
    ws.cell(row=15, column=4, value="Container detail").font = SECTION_FONT
    detail = [
        ("Container ID", "=IF($D13=\"\",\"\",$E13)"),
        ("Barcode", "=IF($D13=\"\",\"\",$D13)"),
        ("Product", "=IF($D13=\"\",\"\",$G13)"),
        ("Batch / Lot", "=IF($D13=\"\",\"\",$H13)"),
        ("Status", "=IF($D13=\"\",\"\",$L13)"),
        ("Location", "=IF($D13=\"\",\"\",$K13)"),
        ("Expiry", "=IF($D13=\"\",\"\",$I13)"),
        ("Expiry class", "=IF($D13=\"\",\"\",$N13)"),
        ("Opened", '=IF($D13="","",IF($M13="","No","Yes"))'),
        ("Duplicate flag", "=IF($D13=\"\",\"\",$O13)"),
    ]
    for i, (label, formula) in enumerate(detail):
        r = 16 + i
        ws.cell(row=r, column=4, value=label).font = BODY_FONT
        ws.cell(row=r, column=5, value=formula)
        ws.cell(row=r, column=5).alignment = CENTER
    _box_region(ws, 16, 4, 25, 5)

    # validation panel
    ws.cell(row=27, column=4, value="Validation & allowed next actions").font = SECTION_FONT
    ws.cell(row=28, column=4, value="Allowed next actions").font = BODY_FONT
    ws.cell(row=28, column=5, value='=IF($D13="","",IF(AND($L13="Available",$I13<>"",$I13<TODAY()),"TakeOpen BLOCKED (expired by date) | Dispose | MarkExpired",IF($L13="Available","TakeOpen | Transfer | Dispose (confirm) | MarkExpired/Damaged/Missing (confirm)",IF($L13="InUse","Return | Dispose (confirm) | MarkExpired/Damaged/Missing (confirm)",IF($L13="Expired","Dispose | Transfer (relocation only)",IF($L13="Damaged","Dispose | Transfer (relocation only)",IF($L13="Disposed","None — terminal (Adjustment only)","Dispose (resolve) | Adjustment (found)")))))))')
    ws.cell(row=28, column=5).font = Font(size=9)
    ws.cell(row=29, column=4, value="Blocking conditions").font = BODY_FONT
    ws.cell(row=29, column=5, value='=IF($O13="DUPLICATE","DUPLICATE BARCODE — resolve before action",IF($P13="UNKNOWN","UNKNOWN BARCODE — receive first",IF($P13="EMPTY","Scan a barcode",IF(AND($L13="Available",$I13<>"",$I13<TODAY()),"EXPIRED BY DATE — TakeOpen blocked; record MarkExpired or Dispose","None"))))')
    ws.cell(row=29, column=5).font = Font(size=9, color="B71C1C")
    _box_region(ws, 28, 4, 29, 5)

    ws.cell(row=31, column=4, value="Workflow (interim manual mode): scan → review card → perform transaction on Containers/Transactions sheets. VBA phase adds atomic commit + scan reset/focus.").font = SMALL_FONT
    ws.cell(row=32, column=4, value="Scanner: USB keyboard wedge, Enter suffix, Code 128 labels.").font = SMALL_FONT


def _build_receiving(wb):
    ws = wb["Receiving"]
    _title(ws, "Receiving", "Guided receiving interface foundation. VBA phase performs the atomic multi-row append.", ncols=10)
    _set_widths(ws, [6, 18, 3, 20, 30, 14, 14, 14, 14, 12])

    ws.cell(row=4, column=2, value="Receiving entry").font = SECTION_FONT
    ws.cell(row=6, column=2, value="Product ID:").font = BODY_FONT
    ws.cell(row=7, column=4, value="").fill = FILL_INPUT
    ws.cell(row=8, column=2, value="Product name:").font = BODY_FONT
    # Use direct $D$7 references (not the rngReceiveProductID named range) so
    # Excel's dynamic-array engine does not inject implicit-intersection on the
    # named range. This mirrors the Scan sheet, which verifiably works in Excel.
    ws.cell(row=8, column=4, value='=IF($D$7="","",IF(COUNTIF(tblProducts[ProductID],$D$7)=1,INDEX(tblProducts[ProductName],MATCH($D$7,tblProducts[ProductID],0)),"UNKNOWN"))')
    ws.cell(row=9, column=2, value="Next Container ID:").font = BODY_FONT
    ws.cell(row=9, column=4, value='=IF($D$7="","",TEXT(MAX(tblContainers[HelperContainerNum])+1,"C000000"))')
    ws.cell(row=9, column=4).fill = FILL_LIGHT
    ws.cell(row=10, column=2, value="Next Barcode:").font = BODY_FONT
    ws.cell(row=10, column=4, value='=IF($D$7="","",TEXT(MAX(tblContainers[HelperBarcodeNum])+1,"0000000"))')
    ws.cell(row=10, column=4).fill = FILL_LIGHT
    ws.cell(row=12, column=2, value="Batch / Lot:").font = BODY_FONT
    ws.cell(row=12, column=4, value="").fill = FILL_INPUT
    ws.cell(row=13, column=2, value="Expiry date:").font = BODY_FONT
    ws.cell(row=13, column=4, value="").fill = FILL_INPUT
    ws.cell(row=14, column=2, value="Retest date:").font = BODY_FONT
    ws.cell(row=14, column=4, value="").fill = FILL_INPUT
    ws.cell(row=15, column=2, value="Location:").font = BODY_FONT
    ws.cell(row=15, column=4, value="").fill = FILL_INPUT
    ws.cell(row=16, column=2, value="Quantity:").font = BODY_FONT
    ws.cell(row=16, column=4, value=1).fill = FILL_INPUT
    ws.cell(row=17, column=2, value="New status:").font = BODY_FONT
    ws.cell(row=17, column=4, value='=DefaultStatusNewContainers').fill = FILL_LIGHT

    ws.cell(row=19, column=2, value="Readiness:").font = BODY_FONT
    ws.cell(row=19, column=4, value='=IF($D$7="","NOT READY — select product",IF($D$12="","NOT READY — enter lot",IF(AND($D$16>=1,$D$16<=999),"READY","NOT READY — quantity")))')
    ws.cell(row=19, column=4).font = Font(bold=True, size=11)
    ws.cell(row=21, column=2, value="Steps (interim manual mode):").font = SECTION_FONT
    steps = [
        "1. Select Product ID (list from tblProducts).",
        "2. Confirm next Container ID and Barcode (formula-generated).",
        "3. Enter lot, expiry (if any), location, quantity.",
        "4. Append one row to tblContainers (ContainerID, Barcode, ProductID, lot, dates, location, Status=Available).",
        "5. Append one Receive row to tblTransactions (snapshot all fields; PreviousStatus=(none)).",
        "6. Repeat per container for batch receiving.",
    ]
    for i, s in enumerate(steps):
        ws.cell(row=22 + i, column=2, value=s).font = SMALL_FONT

    # staging table header (tblReceiveStaging)
    ws.cell(row=30, column=2, value="tblReceiveStaging (formula staging — not authoritative)").font = SECTION_FONT
    rcols = TABLES["tblReceiveStaging"]["columns"]
    hdr = 31
    for i, (name, header, _w) in enumerate(rcols, start=1):
        ws.cell(row=hdr, column=1 + i, value=header)
    _style_header(ws, hdr, len(rcols), start_col=2)
    ws.cell(row=32, column=2, value='=IF($D$7="","",$D$7)')
    ws.cell(row=32, column=3, value='=IF($D$7="","",B8)')
    ws.cell(row=32, column=4, value='=IF($D$7="","",D9)')
    ws.cell(row=32, column=5, value='=IF($D$7="","",D10)')
    ws.cell(row=32, column=6, value='=IF($D$7="","",D12)')
    ws.cell(row=32, column=7, value='=IF($D$7="","",D13)')
    ws.cell(row=32, column=8, value='=IF($D$7="","",D14)')
    ws.cell(row=32, column=9, value='=IF($D$7="","",D15)')
    ws.cell(row=32, column=10, value='=IF($D$7="","",D17)')
    ws.cell(row=32, column=11, value='=IF($D$7="","",D16)')
    ws.cell(row=32, column=12, value='=IF($D$7="","",D19)')
    ws.cell(row=32, column=13, value='=IF($D$7="","",D19)')
    for c in range(2, 14):
        ws.cell(row=32, column=c).border = BOX
        ws.cell(row=32, column=c).alignment = CENTER


def _build_dashboard(wb):
    ws = wb["Dashboard"]
    _title(ws, "Laboratory Inventory — Dashboard", "All figures are formula-derived from source Tables (no manual stock numbers).", ncols=8)
    _set_widths(ws, [6, 30, 14, 12, 14, 16, 16, 16])

    ws.cell(row=4, column=2, value="Key statistics").font = SECTION_FONT
    stats = [
        ("Total active products", '=COUNTIF(tblProducts[Active],TRUE)'),
        # D-018: total available = Status=Available AND (no expiry OR expiry>=TODAY)
        ("Total available containers (usable)", '=COUNTIF(tblContainers[Status],"Available")-COUNTIFS(tblContainers[Status],"Available",tblContainers[ExpiryDate],"<"&TODAY())'),
        ("Containers in use", '=COUNTIF(tblContainers[Status],"InUse")'),
        ("Expired containers (any status)", '=COUNTIF(tblContainers[ExpiryDate],"<"&TODAY())'),
        ("Available expiring ≤30 days", '=COUNTIFS(tblContainers[Status],"Available",tblContainers[ExpiryDate],">="&TODAY(),tblContainers[ExpiryDate],"<="&TODAY()+ExpiryWarningDays30)'),
        ("Available expiring 31–60 days", '=COUNTIFS(tblContainers[Status],"Available",tblContainers[ExpiryDate],">"&TODAY()+ExpiryWarningDays30,tblContainers[ExpiryDate],"<="&TODAY()+ExpiryWarningDays60)'),
        ("Available expiring 61–90 days", '=COUNTIFS(tblContainers[Status],"Available",tblContainers[ExpiryDate],">"&TODAY()+ExpiryWarningDays60,tblContainers[ExpiryDate],"<="&TODAY()+ExpiryWarningDays90)'),
        ("Products out of stock", '=COUNTIFS(tblProducts[Active],TRUE,tblProducts[HelperAvailableStock],0)'),
        ("Products below minimum (reorder)", '=COUNTIFS(tblProducts[Active],TRUE,tblProducts[HelperStockClass],"Reorder")'),
        ("Products low (below target)", '=COUNTIFS(tblProducts[Active],TRUE,tblProducts[HelperStockClass],"Low")'),
        ("Received last 14 days", '=COUNTIFS(tblTransactions[TransactionType],"Receive",tblTransactions[Timestamp],">="&TODAY()-14)'),
        ("TakeOpen last 14 days", '=COUNTIFS(tblTransactions[TransactionType],"TakeOpen",tblTransactions[Timestamp],">="&TODAY()-14)'),
        ("Frequently used (TakeOpen count)", '=COUNTIF(tblTransactions[TransactionType],"TakeOpen")'),
    ]
    for i, (label, formula) in enumerate(stats):
        r = 5 + i
        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=3, value=formula)
        ws.cell(row=r, column=3).alignment = CENTER
    _box_region(ws, 5, 2, 4 + len(stats), 3)

    # low/out-of-stock table (2021-compatible: list every active product with
    # its class; conditional-format colors flag low/out). No dynamic arrays.
    ws.cell(row=20, column=2, value="Product stock status (all active products)").font = SECTION_FONT
    ws.cell(row=21, column=2, value="Product").font = HDR_FONT
    ws.cell(row=21, column=3, value="Available").font = HDR_FONT
    ws.cell(row=21, column=4, value="Min").font = HDR_FONT
    ws.cell(row=21, column=5, value="Target").font = HDR_FONT
    ws.cell(row=21, column=6, value="Class").font = HDR_FONT
    _style_header(ws, 21, 5, start_col=2)
    # pull from tblProducts helper columns in row order (products start row 5)
    for i in range(22, 22 + 6):
        r = i - 22 + 1
        ws.cell(row=i, column=2, value=f'=IF(INDEX(tblProducts[ProductID],{r})="","",INDEX(tblProducts[ProductName],{r}))')
        ws.cell(row=i, column=3, value=f'=IF(INDEX(tblProducts[ProductID],{r})="","",INDEX(tblProducts[HelperAvailableStock],{r}))')
        ws.cell(row=i, column=4, value=f'=IF(INDEX(tblProducts[ProductID],{r})="","",INDEX(tblProducts[MinimumContainerStock],{r}))')
        ws.cell(row=i, column=5, value=f'=IF(INDEX(tblProducts[ProductID],{r})="","",INDEX(tblProducts[TargetContainerStock],{r}))')
        ws.cell(row=i, column=6, value=f'=IF(INDEX(tblProducts[ProductID],{r})="","",INDEX(tblProducts[HelperStockClass],{r}))')
        for c in range(2, 7):
            ws.cell(row=i, column=c).border = BOX
            ws.cell(row=i, column=c).alignment = CENTER
    # color the class column
    for i in range(22, 22 + 6):
        ws.conditional_formatting.add(
            f"F{i}:F{i}",
            FormulaRule(formula=[f'$F{i}="Reorder"'], fill=FILL_WARN, stopIfTrue=False),
        )
        ws.conditional_formatting.add(
            f"F{i}:F{i}",
            FormulaRule(formula=[f'$F{i}="OutOfStock"'], fill=FILL_BAD, stopIfTrue=False),
        )
        ws.conditional_formatting.add(
            f"F{i}:F{i}",
            FormulaRule(formula=[f'$F{i}="OK"'], fill=FILL_OK, stopIfTrue=False),
        )

    ws.cell(row=32, column=2, value="Inventory by category").font = SECTION_FONT
    ws.cell(row=33, column=2, value="Category").font = HDR_FONT
    ws.cell(row=33, column=3, value="Available").font = HDR_FONT
    _style_header(ws, 33, 2, start_col=2)
    cats = ["Pipette Tips", "Tubes", "Solvent", "Reagent", "Consumable", "General"]
    for i, cat in enumerate(cats):
        r = 34 + i
        ws.cell(row=r, column=2, value=cat)
        # sum HelperAvailableStock across products whose Category == cat
        ws.cell(row=r, column=3, value=f'=SUMPRODUCT((tblProducts[Category]="{cat}")*tblProducts[HelperAvailableStock])')
        ws.cell(row=r, column=3).alignment = CENTER
        ws.cell(row=r, column=2).border = BOX
        ws.cell(row=r, column=3).border = BOX

    # frequently used products (Req §18): count TakeOpen transactions per
    # product, list in deterministic product order (2021-compatible INDEX).
    ws.cell(row=41, column=2, value="Most frequently used products (TakeOpen count)").font = SECTION_FONT
    ws.cell(row=42, column=2, value="Product").font = HDR_FONT
    ws.cell(row=42, column=3, value="TakeOpen count").font = HDR_FONT
    _style_header(ws, 42, 2, start_col=2)
    for i in range(6):
        r = 43 + i
        pr = i + 1  # product row offset (products start row 5)
        ws.cell(row=r, column=2, value=f'=IF(INDEX(tblProducts[ProductID],{pr})="","",INDEX(tblProducts[ProductName],{pr}))')
        ws.cell(row=r, column=3, value=f'=IF(INDEX(tblProducts[ProductID],{pr})="","",COUNTIFS(tblTransactions[ProductID],INDEX(tblProducts[ProductID],{pr}),tblTransactions[TransactionType],"TakeOpen"))')
        ws.cell(row=r, column=3).alignment = CENTER
        ws.cell(row=r, column=2).border = BOX
        ws.cell(row=r, column=3).border = BOX

    # inventory by storage location (Req §18): count usable available
    # containers per location (D-018 semantics: exclude expired-by-date).
    ws.cell(row=50, column=2, value="Inventory by storage location (usable available)").font = SECTION_FONT
    ws.cell(row=51, column=2, value="Location").font = HDR_FONT
    ws.cell(row=51, column=3, value="Available").font = HDR_FONT
    _style_header(ws, 51, 2, start_col=2)
    n_locs = 6
    for i in range(n_locs):
        r = 52 + i
        lr = i + 1  # locations start row 5
        ws.cell(row=r, column=2, value=f'=IF(INDEX(tblLocations[StorageLocationID],{lr})="","",INDEX(tblLocations[LocationName],{lr}))')
        ws.cell(row=r, column=3, value=(
            f'=IF(INDEX(tblLocations[StorageLocationID],{lr})="","",'
            f'COUNTIFS(tblContainers[StorageLocationID],INDEX(tblLocations[StorageLocationID],{lr}),tblContainers[Status],"Available")'
            f'-COUNTIFS(tblContainers[StorageLocationID],INDEX(tblLocations[StorageLocationID],{lr}),tblContainers[Status],"Available",tblContainers[ExpiryDate],"<"&TODAY()))'
        ))
        ws.cell(row=r, column=3).alignment = CENTER
        ws.cell(row=r, column=2).border = BOX
        ws.cell(row=r, column=3).border = BOX

    ws.cell(row=60, column=2, value="Compliance boundary: operational inventory control only — not an SDS, EHS, GMP/GLP, or validated LIMS system.").font = SMALL_FONT


# ------------------------------------------------------------------ table/name/validation/CF/protection application
def _find_header_row(ws, header_text):
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == header_text:
            return r
    raise RuntimeError(f"header not found: {header_text}")


def _apply_tables(wb):
    for tname, spec in TABLES.items():
        ws = wb[spec["sheet"]]
        start_col = spec.get("start_col", 1)
        # header row is the row where start_col column equals the first header text
        header_text = spec["columns"][0][1]
        header_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=start_col).value == header_text:
                header_row = r
                break
        if header_row is None:
            raise RuntimeError(f"header not found for {tname} (start col {start_col})")
        ncols = len(spec["columns"])
        last_col = get_column_letter(start_col + ncols - 1)
        # body end: last non-empty row in the block
        end_row = header_row + 1
        while end_row <= ws.max_row and ws.cell(row=end_row, column=start_col).value is not None:
            end_row += 1
        end_row -= 1
        if end_row < header_row:
            end_row = header_row
        ref = f"{get_column_letter(start_col)}{header_row}:{last_col}{end_row}"
        tab = Table(displayName=tname, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False,
            showLastColumn=False, showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(tab)


def _apply_named_ranges(wb):
    for rname, (sheet, addr) in NAMED_RANGES.items():
        if ":" in addr:
            a, b = addr.split(":")
            ref = f"'{sheet}'!${a}:${b}"
        else:
            ref = f"'{sheet}'!${addr}:${addr}"
        wb.defined_names.add(DefinedName(rname, attr_text=ref))
    for rname, (sheet, addr, _vals) in LIST_RANGES.items():
        a, b = addr.split(":")
        ref = f"'{sheet}'!${a}:${b}"
        wb.defined_names.add(DefinedName(rname, attr_text=ref))
    for rname, (sheet, col, _expr) in COLUMN_NAMES.items():
        ref = f"'{sheet}'!${col}$1:${col}$1048576"
        wb.defined_names.add(DefinedName(rname, attr_text=ref))
    # settings value names
    for rname, (sheet, addr) in SETTINGS_NAMES.items():
        ref = f"'{sheet}'!${addr}:${addr}"
        wb.defined_names.add(DefinedName(rname, attr_text=ref))


def _apply_validation(wb):
    products = wb["Products"]
    containers = wb["Containers"]
    transactions = wb["Transactions"]
    locations = wb["Locations"]
    scan = wb["Scan"]
    receiving = wb["Receiving"]

    def add_list(ws, formula, ref, allow_blank=True):
        dv = DataValidation(type="list", formula1=formula, allow_blank=allow_blank)
        dv.error = "Value not in list"
        dv.errorTitle = "Invalid value"
        dv.showErrorMessage = True
        ws.add_data_validation(dv)
        dv.add(ref)

    # Products
    add_list(products, "=lstProductType", "E5:E2000")
    add_list(products, "=lstCategory", "F5:F2000")
    add_list(products, "=lstBool", "R5:R2000")
    add_list(products, "=lstSupplierIDs", "K5:K2000")
    # Containers
    add_list(containers, "=lstStatusList", "I5:I2000")
    add_list(containers, "=lstDisposalReason", "L5:L2000")
    add_list(containers, "=lstProductsProductID", "C5:C2000")
    add_list(containers, "=lstLocationIDs", "H5:H2000")
    # Transactions
    add_list(transactions, "=lstTransactionTypeList", "H5:H4000")
    add_list(transactions, "=lstStatusList", "I5:J4000")
    add_list(transactions, "=lstTransactionReason", "N5:N4000")
    # Locations
    add_list(locations, "=lstLocationType", "C5:C2000")
    add_list(locations, "=lstBool", "E5:E2000")
    # Scan input: custom 7-digit text
    dv = DataValidation(type="custom", formula1='=AND(LEN(rngScanInput)=7,ISNUMBER(VALUE(rngScanInput)))', allow_blank=True)
    dv.error = "Barcode must be exactly 7 digits"
    dv.errorTitle = "Invalid barcode"
    dv.showErrorMessage = True
    scan.add_data_validation(dv)
    dv.add("D7")
    # Receiving: ProductID list, Location list, Quantity whole
    add_list(receiving, "=lstProductsProductID", "D7")
    add_list(receiving, "=lstLocationIDs", "D15")
    add_list(receiving, "=lstStatusList", "D17")
    dv = DataValidation(type="whole", operator="between", formula1="1", formula2="999", allow_blank=True)
    dv.error = "Quantity must be a positive whole number"
    receiving.add_data_validation(dv)
    dv.add("D16")


def _apply_conditional_formatting(wb):
    containers = wb["Containers"]
    status_col = get_column_letter(9)
    rng_status = f"{status_col}5:{status_col}2000"
    containers.conditional_formatting.add(
        rng_status,
        FormulaRule(formula=[f'${status_col}5="Available"'], fill=FILL_OK, stopIfTrue=False),
    )
    containers.conditional_formatting.add(
        rng_status,
        FormulaRule(formula=[f'${status_col}5="Expired"'], fill=FILL_BAD, stopIfTrue=False),
    )
    containers.conditional_formatting.add(
        rng_status,
        FormulaRule(formula=[f'OR(${status_col}5="Damaged",${status_col}5="Missing",${status_col}5="Disposed")'], fill=FILL_GRAY, stopIfTrue=False),
    )
    containers.conditional_formatting.add(
        "B5:B2000",
        FormulaRule(formula=['COUNTIF($B$5:$B$2000,$B5)>1'], fill=FILL_BAD, stopIfTrue=False),
    )
    containers.conditional_formatting.add(
        "E5:E2000",
        FormulaRule(formula=['AND($E5<>"",$E5<TODAY())'], fill=FILL_BAD, stopIfTrue=False),
    )
    containers.conditional_formatting.add(
        "E5:E2000",
        FormulaRule(formula=['AND($E5<>"",$E5>=TODAY(),$E5<=TODAY()+90)'], fill=FILL_WARN, stopIfTrue=False),
    )


def _apply_protection(wb):
    # Design: sheet protection with empty password is a deterrent (D-015).
    # By default every cell is locked; we unlock the data-entry cells so the
    # workbook remains usable. Formula/header/instruction cells stay locked.
    from openpyxl.styles import Protection

    # Lock helper formula cells on Products (cols 20,21 = T,U) and Containers
    # (cols 14,15 = N,O) explicitly.
    products = wb["Products"]
    containers = wb["Containers"]
    for r in range(5, 2001):
        products.cell(row=r, column=20).protection = Protection(locked=True)
        products.cell(row=r, column=21).protection = Protection(locked=True)
        containers.cell(row=r, column=14).protection = Protection(locked=True)
        containers.cell(row=r, column=15).protection = Protection(locked=True)

    # Unlock entry cells (Scan/Receiving inputs + data Table bodies)
    unlock = {
        "Scan": ["D7"],
        "Receiving": ["D7", "D12", "D13", "D14", "D15", "D16", "D17"],
        "Products": ["A5:S2000"],
        "Containers": ["A5:M2000"],
        "Transactions": ["A5:P2000"],
        "Suppliers": ["A5:H2000"],
        "Locations": ["A5:E2000"],
        "Settings": [],
        "Dashboard": [],
    }
    from openpyxl.utils.cell import range_boundaries, coordinate_from_string, column_index_from_string
    for sheet, ranges in unlock.items():
        ws = wb[sheet]
        for rng in ranges:
            if ":" in rng:
                c1, r1, c2, r2 = range_boundaries(rng)
                for rr in range(r1, r2 + 1):
                    for cc in range(c1, c2 + 1):
                        ws.cell(row=rr, column=cc).protection = Protection(locked=False)
            else:
                col, row = coordinate_from_string(rng)
                ws.cell(row=row, column=column_index_from_string(col)).protection = Protection(locked=False)

    # Protect every sheet (empty password)
    for name in SHEET_ORDER:
        ws = wb[name]
        ws.protection.sheet = True
        ws.protection.password = ""
        ws.protection.formatCells = False
        ws.protection.selectLockedCells = True
        ws.protection.selectUnlockedCells = False

    # Workbook structure protection (prevent sheet rename/delete)
    wb.security.lockStructure = True
    wb.security.lockWindows = False


def main():
    out = sys.argv[1] if len(sys.argv) > 1 else os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "workbook", "LabInventory_v0.1.xlsx"
    )
    path = build(out)
    print("WROTE", path)


if __name__ == "__main__":
    main()
